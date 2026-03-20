import re
import tempfile
from typing import Optional
from openpyxl import load_workbook
from io import BytesIO


def parse_timecode(tc: str, fps: float = 29.97) -> Optional[float]:
    if not tc or not isinstance(tc, str):
        return None
    tc = tc.strip()
    if tc.lower() in ("overall", "overall ", ""):
        return None
    if tc == "0000":
        return 0.0
    # Handle range like "00:00:21:19 - 00:00:26:05" → use start
    if " - " in tc:
        tc = tc.split(" - ")[0].strip()
    parts = tc.replace(" ", "").split(":")
    try:
        if len(parts) == 4:  # HH:MM:SS:FF
            h, m, s, f = [int(p) for p in parts]
            return round(h * 3600 + m * 60 + s + f / fps, 2)
        if len(parts) == 3:  # HH:MM:SS
            h, m, s = [int(p) for p in parts]
            return float(h * 3600 + m * 60 + s)
        if len(parts) == 2:  # MM:SS
            m, s = [int(p) for p in parts]
            return float(m * 60 + s)
    except (ValueError, IndexError):
        pass
    return None


def detect_sheet_type(sheet_name: str) -> str:
    name = sheet_name.strip()
    if re.match(r"^ani_.*_FB$", name, re.IGNORECASE):
        return "animation"
    if re.match(r"^FB_", name):
        return "reviewer"
    if re.match(r"^images_.*_FB$", name, re.IGNORECASE):
        return "image"
    if re.match(r"^\d+$", name) or name.startswith("Copy of"):
        return "storyboard"
    return "unknown"


def extract_scene_number(text: str) -> Optional[int]:
    m = re.search(r"#S?(\d+)", text)
    if m:
        return int(m.group(1))
    return None


def extract_reviewer_from_sheet_name(sheet_name: str) -> Optional[str]:
    if sheet_name.startswith("FB_"):
        return sheet_name[3:]
    if "_FB" in sheet_name:
        return sheet_name.replace("_FB", "")
    return sheet_name


def parse_animation_sheet(ws, sheet_name: str, fps: float = 29.97) -> list[dict]:
    items = []
    reviewer = extract_reviewer_from_sheet_name(sheet_name)
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=1):
        if not row or len(row) < 2:
            continue
        tc_raw = str(row[0]) if row[0] else None
        comment = row[1]
        if not comment:
            continue
        comment = str(comment).strip()
        tc_seconds = parse_timecode(tc_raw, fps) if tc_raw else None
        items.append({
            "timecode_raw": tc_raw,
            "timecode_seconds": tc_seconds,
            "scene_number": None,
            "reviewer": reviewer,
            "comment": comment,
            "item_index": i,
        })
    return items


def parse_reviewer_sheet(ws, sheet_name: str, fps: float = 29.97) -> list[dict]:
    return parse_animation_sheet(ws, sheet_name, fps)


def parse_image_sheet(ws, sheet_name: str) -> list[dict]:
    items = []
    reviewer = extract_reviewer_from_sheet_name(sheet_name)
    idx = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        comments = []
        if len(row) >= 2 and row[1]:
            comments.append(str(row[1]).strip())
        if len(row) >= 3 and row[2]:
            comments.append(str(row[2]).strip())
        for comment in comments:
            if not comment:
                continue
            idx += 1
            scene = extract_scene_number(comment)
            items.append({
                "timecode_raw": None,
                "timecode_seconds": None,
                "scene_number": scene,
                "reviewer": reviewer,
                "comment": comment,
                "item_index": idx,
            })
    return items


def parse_storyboard_sheet(ws, sheet_name: str) -> list[dict]:
    items = []
    header_row = None
    for row in ws.iter_rows(min_row=1, max_row=5, values_only=False):
        for cell in row:
            if cell.value and "comment" in str(cell.value).lower():
                header_row = cell.row
                break
        if header_row:
            break
    if not header_row:
        header_row = 3

    headers = []
    for cell in ws[header_row]:
        headers.append(str(cell.value) if cell.value else "")

    comment_cols = []
    for i, h in enumerate(headers):
        if "comment" in h.lower():
            reviewer = h.replace("comment", "").strip().strip("()")
            comment_cols.append((i, reviewer or "unknown"))

    idx = 0
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if not row:
            continue
        scene_num = None
        if row[0] and isinstance(row[0], (int, float)):
            scene_num = int(row[0])
        for col_idx, reviewer in comment_cols:
            if col_idx < len(row) and row[col_idx]:
                comment = str(row[col_idx]).strip()
                if not comment:
                    continue
                idx += 1
                items.append({
                    "timecode_raw": None,
                    "timecode_seconds": None,
                    "scene_number": scene_num,
                    "reviewer": reviewer,
                    "comment": comment,
                    "item_index": idx,
                })
    return items


class NumbersSheetAdapter:
    """Adapts a numbers_parser Table to behave like an openpyxl worksheet for our parsers."""
    def __init__(self, table):
        self.table = table
        self.max_row = table.num_rows
        self.max_column = table.num_cols

    def iter_rows(self, min_row=1, max_row=None, values_only=True, **kwargs):
        if max_row is None:
            max_row = self.table.num_rows
        for r in range(min_row - 1, min(max_row, self.table.num_rows)):
            row_vals = []
            for c in range(self.table.num_cols):
                row_vals.append(self.table.cell(r, c).value)
            if values_only:
                yield tuple(row_vals)
            else:
                yield tuple(_FakeCell(r + 1, c + 1, v) for c, v in enumerate(row_vals))

    def __getitem__(self, key):
        if isinstance(key, int):
            return list(self.iter_rows(min_row=key, max_row=key, values_only=False))[0]
        return None


class _FakeCell:
    def __init__(self, row, col, value):
        self.row = row
        self.column = col
        self.value = value
        self.coordinate = f"{chr(64+col)}{row}" if col <= 26 else f"C{col}R{row}"


def parse_numbers(file_data: bytes, fps: float = 29.97) -> dict[str, list[dict]]:
    from numbers_parser import Document
    with tempfile.NamedTemporaryFile(suffix=".numbers", delete=True) as tmp:
        tmp.write(file_data)
        tmp.flush()
        doc = Document(tmp.name)

    result = {}
    for sheet in doc.sheets:
        sheet_name = sheet.name
        sheet_type = detect_sheet_type(sheet_name)
        for table in sheet.tables:
            ws = NumbersSheetAdapter(table)
            if sheet_type == "animation":
                items = parse_animation_sheet(ws, sheet_name, fps)
            elif sheet_type == "reviewer":
                items = parse_reviewer_sheet(ws, sheet_name, fps)
            elif sheet_type == "image":
                items = parse_image_sheet(ws, sheet_name)
            elif sheet_type == "storyboard":
                items = parse_storyboard_sheet(ws, sheet_name)
            else:
                continue
            if items:
                result[sheet_name] = {
                    "type": sheet_type,
                    "items": items,
                }
    return result


def parse_xlsx(file_data: bytes, fps: float = 29.97) -> dict[str, list[dict]]:
    wb = load_workbook(BytesIO(file_data), data_only=True, read_only=True)
    result = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_type = detect_sheet_type(sheet_name)
        if sheet_type == "animation":
            items = parse_animation_sheet(ws, sheet_name, fps)
        elif sheet_type == "reviewer":
            items = parse_reviewer_sheet(ws, sheet_name, fps)
        elif sheet_type == "image":
            items = parse_image_sheet(ws, sheet_name)
        elif sheet_type == "storyboard":
            items = parse_storyboard_sheet(ws, sheet_name)
        else:
            continue
        if items:
            result[sheet_name] = {
                "type": sheet_type,
                "items": items,
            }
    wb.close()
    return result


def parse_file(file_data: bytes, filename: str, fps: float = 29.97) -> dict[str, list[dict]]:
    if filename.endswith(".numbers"):
        return parse_numbers(file_data, fps)
    return parse_xlsx(file_data, fps)
